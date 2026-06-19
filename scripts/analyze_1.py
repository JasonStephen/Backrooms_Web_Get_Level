from __future__ import annotations

import argparse
import csv
import io
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_DIR = ROOT / "outputs" / "md"
DEFAULT_OUTPUT_DIR = ROOT / "outputs" / "entry_exit_1"

HEADING_RE = re.compile(r"^(#{1,6})\s*(.+?)\s*$")
LINK_RE = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")
IMAGE_RE = re.compile(r"!\[([^\]]*)\]\([^)]+\)")
EMPHASIS_RE = re.compile(r"[*_`~#>]+")
PUNCT_RE = re.compile(r"[\s:：;；,，.。!！?？()\[\]{}\-_/\\|]+")

PLAIN_LEVEL_RE = re.compile(r"(?<![A-Za-z0-9])[Ll]evel\s*(-?\d+(?:\.\d+)*)")
SLUG_LEVEL_RE = re.compile(r"(?<![A-Za-z0-9])level-\d+(?:-\d+)*", re.IGNORECASE)
PATH_LEVEL_RE = re.compile(r"(?:^|/)(?:trimmed:|latest:)?(level-\d+(?:-\d+)*)$", re.IGNORECASE)

ENTRANCE_ALIASES = {
    "入口",
    "入囗",  # variant character
    "如何进入",
    "进入方式",
    "进入方法",
    "进入途径",
    "已记录入口",
    "记录入口",
    "如何抵达",
    "抵达方式",
    "进入",
    "通行",
}

EXIT_ALIASES = {
    "出口",
    "出囗",  # variant character
    "如何离开",
    "离开方式",
    "离开方法",
    "离开途径",
    "已记录出口",
    "记录出口",
    "离开",
    "撤离",
    "撤离点",
    "逃生",
    "逃生地点",
    "逃脱",
    "逃离",
    "出路",
    "退路",
    "返回",
    "通行",
}

COMBINED_ALIASES = {
    "入口与出口",
    "入口和出口",
    "出入口",
    "入口出口",
    "入囗与出囗",  # variant character
    "已记录入口与出口",
    "已记录入口和出口",
    "记录入口与出口",
    "记录入口和出口",
    "出入端口",
    "通行",
    "进路与出路",
    "入口及撤离点",
    "入口和逃生地点",
    "入口与撤离",
}

CORE_ENTRANCE_ALIAS = "入口"
CORE_EXIT_ALIAS = "出口"
CORE_COMBINED_ALIASES = {"入口与出口", "入口和出口", "出入口", "入口出口"}

ENTRY_EXIT_SIGNAL_RE = re.compile(r"(入口|出口|进入|离开|逃离|切出|返回|回到|通往|通过|经由|穿过)")

ENT_METHOD_RE = re.compile(r"(进入.*(?:方法|方式|Level|层级)|入口.*(?:位于|在)|可通过(?:.*进入|.*抵达)|可以从(?:.*进入|.*抵达)|经由.*进入|进入.*Level)")
EXIT_METHOD_RE = re.compile(r"(离开.*(?:方法|方式|Level|层级|此地|此处)|出口.*(?:位于|在)|可通过.*离开|可以从.*离开|通往.*Level|切出.*到|返回.*到|穿过.*门|通向.*层级)")

LEAK_NEGATIVE_RE = re.compile(r"(不得而知|无从知晓|没有.*出口|未.*发现.*出口|没有提到出口|不知道.*如何|不明确|未知|尚不可知|从未.*发现|没有记录|无.*记录|从未|未曾|尚无|没有已知|不.*清楚|无法.*(?:离开|进入|找到))")

ENT_KEY = re.compile(r"(入口|出口|进入|离开|逃离|切出|切入|返回|通往|通行|抵达|转移|穿越)")

BOLD_PSEUDO_RE = re.compile(r"^\s*\*\*(.+?)\*\*\s*$")

INLINE_HEADING_RE = re.compile(r"(?<!\n)(#{1,4})\s+([^\n]{1,80}?)(?=\s*$|(?=\s*[-*]))")  # mid-line headings

SENTENCE_SPLIT_RE = re.compile(r"[。！？；\n]")

NON_EE_HEADING_KEYWORDS = {"作者", "授权", "脚注", "来源", "图源", "译者", "翻译", "版权", "协议", "附件", "附录", "投票", "评论"}

NARRATIVE_EXIT_PATTERNS = (
    r"离开了房间",
    r"离开了这里",
    r"离开家",
    r"再也没有.*离开",
    r"别.*离开",
    r"不要.*离开",
    r"未曾.*离开",
    r"从来没有离开过",
    r"不会.*离开",
    r"离开了柜台",
    r"我爱你",
    r"答应我",
    r"一个吻",
    r"两颗心",
    r"喘息着说道",
)

ENTRANCE_KEYWORDS = (
    "进入",
    "进入到",
    "抵达",
    "通往本层级",
    "进入本层级",
    "入口",
    "可经由",
    "可以从",
    "通过",
)

EXIT_KEYWORDS = (
    "离开",
    "逃离",
    "出去",
    "通往",
    "出口",
    "返回",
    "回到",
    "切出",
    "原路返回",
)


@dataclass
class Section:
    title: str
    level: int
    lines: list[str] = field(default_factory=list)
    children: list["Section"] = field(default_factory=list)


def strip_front_matter(text: str) -> str:
    lines = text.splitlines()
    if len(lines) >= 3 and lines[0].strip() == "---":
        for index in range(1, len(lines)):
            if lines[index].strip() == "---":
                return "\n".join(lines[index + 1 :])
    return text


def normalize_heading(text: str) -> str:
    text = LINK_RE.sub(r"\1", text)
    text = EMPHASIS_RE.sub("", text)
    text = text.strip()
    text = PUNCT_RE.sub("", text)
    text = text.replace("和", "与")
    return text


def parse_markdown_sections(text: str) -> Section:
    root = Section(title="", level=0)
    stack: list[Section] = [root]

    for raw_line in strip_front_matter(text).splitlines():
        line = raw_line.rstrip()
        heading_match = HEADING_RE.match(line)
        if heading_match:
            hashes, title = heading_match.groups()
            level = len(hashes)
            section = Section(title=title.strip(), level=level)
            while stack and stack[-1].level >= level:
                stack.pop()
            stack[-1].children.append(section)
            stack.append(section)
            continue

        stack[-1].lines.append(line)

    return root


def collect_text_lines(section: Section) -> list[str]:
    lines = list(section.lines)
    for child in section.children:
        lines.extend(collect_text_lines(child))
    return lines


def clean_block(block: str) -> str:
    block = IMAGE_RE.sub(r"\1", block)
    block = LINK_RE.sub(r"\1", block)
    block = EMPHASIS_RE.sub("", block)
    block = re.sub(r"\s+", " ", block).strip()
    return block


def split_blocks(lines: Iterable[str]) -> list[str]:
    blocks: list[str] = []
    buffer: list[str] = []

    def flush() -> None:
        if not buffer:
            return
        text = clean_block(" ".join(buffer))
        if text:
            blocks.append(text)
        buffer.clear()

    for line in lines:
        stripped = line.strip()
        if not stripped:
            flush()
            continue

        if stripped.startswith(("-", "*")) or re.match(r"^\d+\.", stripped):
            flush()
            text = clean_block(re.sub(r"^(-|\*|\d+\.)\s*", "", stripped))
            if text:
                blocks.append(text)
            continue

        buffer.append(stripped)

    flush()
    return blocks


def classify_block(block: str) -> str | None:
    if not ENTRY_EXIT_SIGNAL_RE.search(block):
        return None

    entrance_score = sum(keyword in block for keyword in ENTRANCE_KEYWORDS)
    exit_score = sum(keyword in block for keyword in EXIT_KEYWORDS)

    if "离开" in block or "逃离" in block or "原路返回" in block:
        exit_score += 2
    if "进入" in block and "离开" not in block and "逃离" not in block:
        entrance_score += 1

    if entrance_score == 0 and exit_score == 0:
        return None
    if entrance_score > exit_score:
        return "entrances"
    if exit_score > entrance_score:
        return "exits"

    if "进入" in block and "离开" not in block:
        return "entrances"
    if "离开" in block and "进入" not in block:
        return "exits"
    return None


def find_sections(section: Section, normalized_titles: set[str]) -> list[Section]:
    matches: list[Section] = []
    for child in section.children:
        norm = normalize_heading(child.title)
        if norm in normalized_titles:
            matches.append(child)
        elif any(alias in norm for alias in normalized_titles):
            matches.append(child)
        matches.extend(find_sections(child, normalized_titles))
    return matches


def unique_blocks(blocks: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for block in blocks:
        if block and block not in seen:
            seen.add(block)
            result.append(block)
    return result


def _matches_alias(norm: str, aliases: set[str]) -> bool:
    if norm in aliases:
        return True
    return any(alias in norm for alias in aliases)


def _section_has_entry_exit_signal(section: Section) -> bool:
    all_text = " ".join(collect_text_lines(section))
    return bool(ENTRY_EXIT_SIGNAL_RE.search(all_text))


def extract_from_combined_section(section: Section) -> tuple[list[str], list[str]]:
    entrances: list[str] = []
    exits: list[str] = []

    # Find marker section indices among children
    entrance_indices: list[int] = []
    exit_indices: list[int] = []
    for i, child in enumerate(section.children):
        norm = normalize_heading(child.title)
        if _matches_alias(norm, ENTRANCE_ALIASES) and not _matches_alias(norm, COMBINED_ALIASES):
            entrance_indices.append(i)
        if _matches_alias(norm, EXIT_ALIASES) and not _matches_alias(norm, COMBINED_ALIASES):
            exit_indices.append(i)

    section_blocks = split_blocks(section.lines)

    if entrance_indices and exit_indices:
        first_ent = entrance_indices[0]
        first_exit = exit_indices[0]

        # Entrance: all children between entrance and exit markers (skip noise)
        entrances.extend(section_blocks)
        for child in section.children[first_ent:first_exit]:
            if not _is_noise_heading(child.title):
                entrances.extend(split_blocks(collect_text_lines(child)))

        # Exit: all children from exit marker onwards (skip noise)
        for child in section.children[first_exit:]:
            if not _is_noise_heading(child.title):
                exits.extend(split_blocks(collect_text_lines(child)))

        return unique_blocks(entrances), unique_blocks(exits)

    if entrance_indices:
        first_ent = entrance_indices[0]
        entrances.extend(section_blocks)
        for child in section.children[first_ent:]:
            if _section_has_entry_exit_signal(child):
                entrances.extend(split_blocks(collect_text_lines(child)))
        return unique_blocks(entrances), unique_blocks(exits)

    if exit_indices:
        first_exit = exit_indices[0]
        exits.extend(section_blocks)
        for child in section.children[first_exit:]:
            if _section_has_entry_exit_signal(child):
                exits.extend(split_blocks(collect_text_lines(child)))
        return unique_blocks(entrances), unique_blocks(exits)

    # No child markers found — classify section.lines text blocks
    current_bucket: str | None = None
    for block in section_blocks:
        normalized_block = normalize_heading(block)
        if _matches_alias(normalized_block, ENTRANCE_ALIASES) and not _matches_alias(normalized_block, COMBINED_ALIASES):
            current_bucket = "entrances"
            continue
        if _matches_alias(normalized_block, EXIT_ALIASES) and not _matches_alias(normalized_block, COMBINED_ALIASES):
            current_bucket = "exits"
            continue

        if current_bucket == "entrances":
            entrances.append(block)
            continue
        if current_bucket == "exits":
            exits.append(block)
            continue

        bucket = classify_block(block)
        if bucket == "entrances":
            entrances.append(block)
        elif bucket == "exits":
            exits.append(block)

        # For long blocks, also try sentence-level split to catch mixed content
        if len(block) > 120:
            sentences = [s.strip() for s in SENTENCE_SPLIT_RE.split(block) if s.strip() and len(s.strip()) > 10]
            for sent in sentences:
                sent_bucket = classify_block(sent)
                if sent_bucket == "entrances":
                    entrances.append(sent)
                elif sent_bucket == "exits":
                    exits.append(sent)

    return unique_blocks(entrances), unique_blocks(exits)


def _scan_body_for_ee(tree: Section) -> tuple[list[str], list[str]]:
    blocks = split_blocks(collect_text_lines(tree))
    entrances: list[str] = []
    exits: list[str] = []
    for block in blocks:
        if _is_narrative_noise(block):
            continue
        bucket = classify_block(block)
        if bucket == "entrances":
            entrances.append(block)
        elif bucket == "exits":
            exits.append(block)

        if len(block) > 120:
            sentences = [s.strip() for s in SENTENCE_SPLIT_RE.split(block) if s.strip() and len(s.strip()) > 10]
            for sent in sentences:
                if _is_narrative_noise(sent):
                    continue
                sent_bucket = classify_block(sent)
                if sent_bucket == "entrances":
                    entrances.append(sent)
                elif sent_bucket == "exits":
                    exits.append(sent)
    return entrances, exits


def _extract_from_standalone_section(section: Section, is_entrance: bool) -> tuple[list[str], list[str]]:
    """Process a standalone section that may contain both entrance and exit children."""
    entrances: list[str] = []
    exits: list[str] = []

    # Find opposing markers among children
    other_indices: list[int] = []
    search_aliases = EXIT_ALIASES if is_entrance else ENTRANCE_ALIASES
    for i, child in enumerate(section.children):
        norm = normalize_heading(child.title)
        if _matches_alias(norm, search_aliases) and not _matches_alias(norm, COMBINED_ALIASES):
            other_indices.append(i)

    # Section intro lines belong to the primary type
    blocks = split_blocks(section.lines)
    (entrances if is_entrance else exits).extend(blocks)

    if other_indices:
        split_at = other_indices[0]
        before = section.children[:split_at]
        after = section.children[split_at:]

        # Before split → primary type
        src_primary = entrances if is_entrance else exits
        for child in before:
            if not _is_noise_heading(child.title) and _section_has_entry_exit_signal(child):
                src_primary.extend(split_blocks(collect_text_lines(child)))

        # After split → opposite type
        src_other = exits if is_entrance else entrances
        for child in after:
            if not _is_noise_heading(child.title):
                src_other.extend(split_blocks(collect_text_lines(child)))
    else:
        src = entrances if is_entrance else exits
        for child in section.children:
            if _section_has_entry_exit_signal(child):
                src.extend(split_blocks(collect_text_lines(child)))

    return entrances, exits


def extract_entry_exit(text: str) -> dict[str, object]:
    tree = parse_markdown_sections(text)
    entrances: list[str] = []
    exits: list[str] = []

    combined_sections = find_sections(tree, COMBINED_ALIASES)
    if combined_sections:
        combined_entrances, combined_exits = extract_from_combined_section(combined_sections[-1])
        entrances.extend(combined_entrances)
        exits.extend(combined_exits)

    standalone_entrances = find_sections(tree, ENTRANCE_ALIASES)
    standalone_exits = find_sections(tree, EXIT_ALIASES)

    if standalone_entrances:
        st_ent, st_exit = _extract_from_standalone_section(standalone_entrances[-1], is_entrance=True)
        entrances.extend(st_ent)
        exits.extend(st_exit)

    if standalone_exits:
        st_ent, st_exit = _extract_from_standalone_section(standalone_exits[-1], is_entrance=False)
        entrances.extend(st_ent)
        exits.extend(st_exit)

    entrances = unique_blocks(entrances)
    exits = unique_blocks(exits)

    if not entrances and not exits:
        body_entrances, body_exits = _scan_body_for_ee(tree)
        if body_entrances or body_exits:
            return {
                "entrances": unique_blocks(body_entrances),
                "exits": unique_blocks(body_exits),
                "confidence": "low",
            }
        return {
            "entrances": [],
            "exits": [],
            "confidence": "none",
        }

    return {
        "entrances": entrances,
        "exits": exits,
        "confidence": "high",
    }


def logical_id_from_filename(path: Path) -> str:
    return path.name.removesuffix(".body.md")


def _natural_sort_key(path: Path) -> tuple[list[int], str]:
    """Sort key: numeric parts first, then raw string for tie-breaking."""
    logical_id = path.name.removesuffix(".body.md")
    parts = logical_id.split("-")
    nums: list[int] = []
    for p in parts:
        try:
            nums.append(int(p))
        except ValueError:
            nums.append(10**9)  # non-numeric like "sanctum" goes to end
    return (nums, logical_id)


def normalize_level_token(token: str) -> str:
    value = token.lower().strip()
    if value.startswith("level-"):
        return value
    if value.startswith("level "):
        value = value.replace("level ", "level-", 1)
    if value.startswith("level"):
        value = value.replace("level", "level-", 1)
    value = value.replace(".", "-")
    value = re.sub(r"-{2,}", "-", value)
    return value


def extract_level_ids_from_text(text: str) -> list[str]:
    level_ids: list[str] = []

    for match in SLUG_LEVEL_RE.finditer(text):
        level_ids.append(match.group(0).lower())

    for match in PLAIN_LEVEL_RE.finditer(text):
        number_part = match.group(1).replace(".", "-")
        level_ids.append(f"level-{number_part}".lower())

    return unique_blocks(level_ids)


def extract_level_ids_from_block(block: str, current_level_id: str) -> list[str]:
    level_ids = [level_id for level_id in extract_level_ids_from_text(block) if level_id != current_level_id]
    return unique_blocks(level_ids)


def _determine_status(has_desc: bool, is_leak: bool) -> str:
    if has_desc:
        return "已捕获"
    if is_leak:
        return "漏网"
    return "确认无"


def _is_noise_heading(title: str) -> bool:
    norm = normalize_heading(title)
    return any(kw in norm for kw in NON_EE_HEADING_KEYWORDS)


def _is_narrative_noise(block: str) -> bool:
    for pat in NARRATIVE_EXIT_PATTERNS:
        if re.search(pat, block):
            return True
    return False


def _check_body_has_leak(text: str, method_re: re.Pattern) -> bool:
    lines = text.splitlines()
    fm_end = 0
    if len(lines) >= 3 and lines[0].strip() == "---":
        for i in range(1, len(lines)):
            if lines[i].strip() == "---":
                fm_end = i + 1
                break
    body = "\n".join(lines[fm_end:])
    # Must match a method pattern AND not be a negative statement
    if not method_re.search(body):
        return False
    # Check if the matched text is negated (e.g. "方法不得而知")
    for m in method_re.finditer(body):
        start = max(0, m.start() - 10)
        end = min(len(body), m.end() + 30)
        context = body[start:end]
        if not LEAK_NEGATIVE_RE.search(context):
            return True
    return False


def _has_bold_pseudo_heading(text: str, target: str) -> bool:
    """Check if article has bold-text pseudo-headings for entrance/exit."""
    lines = text.splitlines()
    fm_end = 0
    if len(lines) >= 3 and lines[0].strip() == "---":
        for i in range(1, len(lines)):
            if lines[i].strip() == "---":
                fm_end = i + 1
                break
    for line in lines[fm_end:]:
        m = BOLD_PSEUDO_RE.match(line)
        if m:
            norm = normalize_heading(m.group(1))
            if target == "entrance":
                if _matches_alias(norm, ENTRANCE_ALIASES) and not _matches_alias(norm, COMBINED_ALIASES):
                    return True
            elif target == "exit":
                if _matches_alias(norm, EXIT_ALIASES) and not _matches_alias(norm, COMBINED_ALIASES):
                    return True
            if _matches_alias(norm, COMBINED_ALIASES):
                return True
    return False


def analyze_file(path: Path) -> dict[str, object]:
    text = path.read_text(encoding="utf-8")
    extracted = extract_entry_exit(text)
    logical_id = logical_id_from_filename(path)

    entrance_level_ids = unique_blocks(
        level_id
        for block in extracted["entrances"]
        for level_id in extract_level_ids_from_block(block, logical_id)
    )
    exit_level_ids = unique_blocks(
        level_id
        for block in extracted["exits"]
        for level_id in extract_level_ids_from_block(block, logical_id)
    )

    has_ent = bool(extracted["entrances"])
    has_exit = bool(extracted["exits"])
    ent_leak = not has_ent and _check_body_has_leak(text, ENT_METHOD_RE)
    exit_leak = not has_exit and _check_body_has_leak(text, EXIT_METHOD_RE)
    ent_bold = _has_bold_pseudo_heading(text, "entrance")
    exit_bold = _has_bold_pseudo_heading(text, "exit")

    return {
        "file": path.name,
        "logical_id": logical_id,
        "entrance_count": len(extracted["entrances"]),
        "exit_count": len(extracted["exits"]),
        "entrances": extracted["entrances"],
        "exits": extracted["exits"],
        "entrance_level_ids": entrance_level_ids,
        "exit_level_ids": exit_level_ids,
        "missing_entrance_level_ids": has_ent and not entrance_level_ids,
        "missing_exit_level_ids": has_exit and not exit_level_ids,
        "confidence": extracted.get("confidence", "none"),
        "ent_status": _determine_status(has_ent, ent_leak),
        "exit_status": _determine_status(has_exit, exit_leak),
        "ent_bold": ent_bold,
        "exit_bold": exit_bold,
    }


def build_description_output(results: list[dict[str, object]]) -> dict[str, object]:
    return {
        "file_count": len(results),
        "with_entrances": sum(bool(item["entrances"]) for item in results),
        "with_exits": sum(bool(item["exits"]) for item in results),
        "results": results,
    }


def build_level_output(results: list[dict[str, object]]) -> dict[str, object]:
    compact_results = []
    for item in results:
        compact_results.append(
            {
                "file": item["file"],
                "logical_id": item["logical_id"],
                "entrance_level_ids": item["entrance_level_ids"],
                "exit_level_ids": item["exit_level_ids"],
            }
        )

    return {
        "file_count": len(compact_results),
        "with_entrance_level_ids": sum(bool(item["entrance_level_ids"]) for item in compact_results),
        "with_exit_level_ids": sum(bool(item["exit_level_ids"]) for item in compact_results),
        "results": compact_results,
    }


def build_missing_level_info_text(results: list[dict[str, object]]) -> str:
    lines = ["Missing entrance/exit level ids", ""]

    for item in results:
        missing_entry = item["missing_entrance_level_ids"]
        missing_exit = item["missing_exit_level_ids"]
        if not missing_entry and not missing_exit:
            continue

        lines.append(f"{item['logical_id']} ({item['file']})")
        if missing_entry:
            lines.append("  [entrances]")
            for block in item["entrances"]:
                lines.append(f"  - {block}")
        if missing_exit:
            lines.append("  [exits]")
            for block in item["exits"]:
                lines.append(f"  - {block}")
        lines.append("")

    if len(lines) == 2:
        lines.append("None")

    return "\n".join(lines)


def build_missing_descriptions_text(results: list[dict[str, object]]) -> str:
    lines = ["Missing entrance/exit descriptions", ""]

    for item in results:
        missing_entrances = not item["entrances"]
        missing_exits = not item["exits"]
        if not missing_entrances and not missing_exits:
            continue

        lines.append(f"{item['logical_id']} ({item['file']})")
        if missing_entrances:
            lines.append("  - missing entrances")
        if missing_exits:
            lines.append("  - missing exits")
        lines.append("")

    if len(lines) == 2:
        lines.append("None")

    return "\n".join(lines)


def build_status_report_csv(results: list[dict[str, object]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["logical_id", "入口描述", "入口层级", "出口描述", "出口层级", "入口标题", "出口标题", "入口状态", "出口状态", "置信度"])

    for item in results:
        has_entrance_description = bool(item["entrances"])
        has_entrance_level_ids = bool(item["entrance_level_ids"])
        has_exit_description = bool(item["exits"])
        has_exit_level_ids = bool(item["exit_level_ids"])

        confidence = item.get("confidence", "none")
        ent_status = item.get("ent_status", "")

        # Title hit for CSV
        if confidence == "high" and has_entrance_description:
            ent_title = "Y"
        elif item.get("ent_bold"):
            ent_title = "Y"
        elif ent_status == "漏网":
            ent_title = "LEAK"
        elif has_entrance_description:
            ent_title = "N"
        else:
            ent_title = "N"

        if confidence == "high" and has_exit_description:
            exit_title = "Y"
        elif item.get("exit_bold"):
            exit_title = "Y"
        elif item.get("exit_status", "") == "漏网":
            exit_title = "LEAK"
        elif has_exit_description:
            exit_title = "N"
        else:
            exit_title = "N"

        writer.writerow([
            item["logical_id"],
            has_entrance_description,
            has_entrance_level_ids,
            has_exit_description,
            has_exit_level_ids,
            ent_title,
            exit_title,
            ent_status,
            item.get("exit_status", ""),
            confidence,
        ])

    return output.getvalue()


def build_status_report_xlsx(results: list[dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "出入口状态"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    green_font = Font(color="006100", size=11)
    red_font = Font(color="9C0006", size=11)
    yellow_font = Font(color="9C6500", size=11)
    grey_font = Font(color="595959", size=11)

    headers = ["logical_id", "入口描述", "入口层级", "出口描述", "出口层级", "入口标题", "出口标题", "入口状态", "出口状态", "置信度"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx, item in enumerate(results, 2):
        has_entrance_description = bool(item["entrances"])
        has_entrance_level_ids = bool(item["entrance_level_ids"])
        has_exit_description = bool(item["exits"])
        has_exit_level_ids = bool(item["exit_level_ids"])
        confidence = item.get("confidence", "none")
        confidence_map = {"high": "高", "low": "低", "none": "无"}

        ent_status = item.get("ent_status", "")
        exit_status = item.get("exit_status", "")

        # Title hit: ✔=heading found, ⚠=leak, ❌=not found
        if confidence == "high" and has_entrance_description:
            ent_title = "✔"
        elif item.get("ent_bold"):
            ent_title = "✔"
        elif ent_status == "漏网":
            ent_title = "⚠"
        elif has_entrance_description:
            ent_title = "❌"
        else:
            ent_title = "❌"

        if confidence == "high" and has_exit_description:
            exit_title = "✔"
        elif item.get("exit_bold"):
            exit_title = "✔"
        elif exit_status == "漏网":
            exit_title = "⚠"
        elif has_exit_description:
            exit_title = "❌"
        else:
            exit_title = "❌"

        values = [
            item["logical_id"],
            "✔ 有" if has_entrance_description else "✘ 无",
            "✔ 有" if has_entrance_level_ids else "✘ 无",
            "✔ 有" if has_exit_description else "✘ 无",
            "✔ 有" if has_exit_level_ids else "✘ 无",
            ent_title,
            exit_title,
            ent_status,
            exit_status,
            confidence_map.get(confidence, confidence),
        ]

        status_fill = {"已捕获": green_fill, "漏网": yellow_fill, "确认无": grey_fill}
        status_font = {"已捕获": green_font, "漏网": yellow_font, "确认无": grey_font}
        title_fill2 = {"✔": green_fill, "❌": red_fill, "⚠": yellow_fill}
        title_font2 = {"✔": green_font, "❌": red_font, "⚠": yellow_font}

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center_align
            if col == 1:
                cell.font = Font(size=11)
            elif col in (6, 7):
                tf = title_fill2.get(value)
                tfn = title_font2.get(value)
                if tf:
                    cell.fill = tf
                    cell.font = tfn
            elif col in (8, 9):
                sf = status_fill.get(value)
                sfn = status_font.get(value)
                if sf:
                    cell.fill = sf
                    cell.font = sfn
                else:
                    cell.fill = grey_fill
                    cell.font = grey_font
            elif col == 10:
                if value == "高":
                    cell.fill = green_fill
                    cell.font = green_font
                elif value == "低":
                    cell.fill = yellow_fill
                    cell.font = yellow_font
                else:
                    cell.fill = grey_fill
                    cell.font = grey_font
            elif "✔" in str(value):
                cell.fill = green_fill
                cell.font = green_font
            else:
                cell.fill = red_fill
                cell.font = red_font

    col_widths = [24, 12, 12, 12, 12, 12, 12, 12, 12, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = f"A1:J{len(results) + 1}"
    ws.freeze_panes = "A2"

    # ----- Sheet 2: 统计分析 -----
    ws2 = wb.create_sheet("统计分析")

    total = len(results)
    ent_desc = sum(1 for r in results if r["entrances"])
    ent_desc_id = sum(1 for r in results if r["entrances"] and r["entrance_level_ids"])
    exit_desc = sum(1 for r in results if r["exits"])
    exit_desc_id = sum(1 for r in results if r["exits"] and r["exit_level_ids"])
    both_desc = sum(1 for r in results if r["entrances"] and r["exits"])
    both_desc_id = sum(1 for r in results if r["entrances"] and r["entrance_level_ids"] and r["exits"] and r["exit_level_ids"])
    high_count = sum(1 for r in results if r.get("confidence") == "high")
    low_count = sum(1 for r in results if r.get("confidence") == "low")
    none_count = sum(1 for r in results if r.get("confidence") == "none")

    stats = [
        ("总文章数", total, ""),
        ("", "", ""),
        ("指标", "命中数", "成功率"),
        ("入口描述文本", ent_desc, f"{ent_desc/total*100:.1f}%"),
        ("入口描述文本 + 层级ID", ent_desc_id, f"{ent_desc_id/total*100:.1f}%"),
        ("出口描述文本", exit_desc, f"{exit_desc/total*100:.1f}%"),
        ("出口描述文本 + 层级ID", exit_desc_id, f"{exit_desc_id/total*100:.1f}%"),
        ("入口 + 出口描述文本", both_desc, f"{both_desc/total*100:.1f}%"),
        ("入口 + 出口描述文本 + 双ID", both_desc_id, f"{both_desc_id/total*100:.1f}%"),
        ("", "", ""),
        ("置信度分布", "数量", "占比"),
        ("高（标题命中）", high_count, f"{high_count/total*100:.1f}%"),
        ("低（正文推断）", low_count, f"{low_count/total*100:.1f}%"),
        ("无（未命中）", none_count, f"{none_count/total*100:.1f}%"),
    ]

    title_font = Font(bold=True, size=14, color="4472C4")
    header_fill2 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white2 = Font(bold=True, size=11, color="FFFFFF")
    percent_font_green = Font(bold=True, size=11, color="006100")
    percent_font_yellow = Font(bold=True, size=11, color="9C6500")
    percent_font_red = Font(bold=True, size=11, color="9C0006")

    ws2.merge_cells("A1:C1")
    ws2.cell(row=1, column=1, value="出入口信息获取成功率统计").font = title_font

    for row_idx, (label, count, rate) in enumerate(stats, 2):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=count if count != "" else "")
        cell_c = ws2.cell(row=row_idx, column=3, value=rate)

        if label == "指标" or label == "置信度分布":
            for c in [cell_a, cell_b, cell_c]:
                c.font = header_font_white2
                c.fill = header_fill2
                c.alignment = center_align
        elif label and not label.startswith("总"):
            cell_a.alignment = Alignment(horizontal="right")
            cell_b.alignment = center_align
            cell_c.alignment = center_align

            pct = float(rate.strip("%"))
            if pct >= 60:
                cell_c.font = percent_font_green
            elif pct >= 45:
                cell_c.font = percent_font_yellow
            else:
                cell_c.font = percent_font_red

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    # ----- Sheet 3: 排除确认无的命中率 -----
    ws3 = wb.create_sheet("有效命中率")

    ent_none_count = sum(1 for r in results if r.get("ent_status") == "确认无")
    exit_none_count = sum(1 for r in results if r.get("exit_status") == "确认无")
    both_none_count = sum(1 for r in results if r.get("ent_status") == "确认无" and r.get("exit_status") == "确认无")

    ent_adj = total - ent_none_count
    exit_adj = total - exit_none_count
    both_adj = total - both_none_count

    adj_stats = [
        ("调整依据", "", ""),
        ("确认无入口的层级数", ent_none_count, "已从入口统计中排除"),
        ("确认无出口的层级数", exit_none_count, "已从出口统计中排除"),
        ("确认双无的层级数", both_none_count, "已从双统计中排除"),
        ("", "", ""),
        ("指标（排除确认无后）", "命中数", "调整后成功率"),
        ("入口描述文本", ent_desc, f"{ent_desc/ent_adj*100:.1f}%"),
        ("入口描述文本 + 层级ID", ent_desc_id, f"{ent_desc_id/ent_adj*100:.1f}%"),
        ("出口描述文本", exit_desc, f"{exit_desc/exit_adj*100:.1f}%"),
        ("出口描述文本 + 层级ID", exit_desc_id, f"{exit_desc_id/exit_adj*100:.1f}%"),
        ("入口 + 出口描述文本", both_desc, f"{both_desc/both_adj*100:.1f}%"),
        ("入口 + 出口描述文本 + 双ID", both_desc_id, f"{both_desc_id/both_adj*100:.1f}%"),
        ("", "", ""),
        ("对照：原始命中率", "", ""),
        ("入口描述文本（原始）", ent_desc, f"{ent_desc/total*100:.1f}%"),
        ("出口描述文本（原始）", exit_desc, f"{exit_desc/total*100:.1f}%"),
        ("入口+出口描述（原始）", both_desc, f"{both_desc/total*100:.1f}%"),
    ]

    ws3.merge_cells("A1:C1")
    ws3.cell(row=1, column=1, value="有效命中率（排除确认无出入口的层级）").font = title_font

    for row_idx, (label, count, rate) in enumerate(adj_stats, 2):
        cell_a = ws3.cell(row=row_idx, column=1, value=label)
        cell_b = ws3.cell(row=row_idx, column=2, value=count if count != "" else "")
        cell_c = ws3.cell(row=row_idx, column=3, value=rate)

        if label in ("调整依据", "指标（排除确认无后）", "对照：原始命中率"):
            for c in [cell_a, cell_b, cell_c]:
                c.font = header_font_white2
                c.fill = header_fill2
                c.alignment = center_align
        elif label and not label.startswith("确认"):
            cell_a.alignment = Alignment(horizontal="right")
            cell_b.alignment = center_align
            cell_c.alignment = center_align

            pct_str = rate.strip("%")
            if pct_str:
                pct = float(pct_str)
                if pct >= 90:
                    cell_c.font = percent_font_green
                elif pct >= 70:
                    cell_c.font = percent_font_yellow
                else:
                    cell_c.font = percent_font_red

    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 18

    wb.save(output_path)


def write_outputs(results: list[dict[str, object]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    description_path = output_dir / "entry_exit_descriptions.json"
    level_path = output_dir / "entry_exit_level_ids.json"
    status_csv_path = output_dir / "entry_exit_status.csv"
    status_xlsx_path = output_dir / "entry_exit_status.xlsx"

    description_path.write_text(
        json.dumps(build_description_output(results), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    level_path.write_text(
        json.dumps(build_level_output(results), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    status_csv_path.write_text(build_status_report_csv(results), encoding="utf-8")
    build_status_report_xlsx(results, status_xlsx_path)

    for stale_txt in [
        output_dir / "missing_level_ids.txt",
        output_dir / "missing_entry_exit_descriptions.txt",
    ]:
        if stale_txt.exists():
            stale_txt.unlink()

    print(f"Wrote description output to {description_path}")
    print(f"Wrote level-id output to {level_path}")
    print(f"Wrote status report to {status_csv_path}")
    print(f"Wrote status xlsx to {status_xlsx_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract Backrooms entry and exit descriptions and level ids.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR, help="Directory containing markdown files.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for generated output files.")
    parser.add_argument("--pattern", default="*.body.md", help="Glob pattern used to select files.")
    parser.add_argument("--limit", type=int, help="Optional limit for debugging.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    files = sorted(args.input_dir.glob(args.pattern), key=_natural_sort_key)
    if args.limit is not None:
        files = files[: args.limit]

    results = [analyze_file(path) for path in files]
    write_outputs(results, args.output_dir)


if __name__ == "__main__":
    main()
