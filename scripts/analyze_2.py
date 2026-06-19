from __future__ import annotations

import argparse
import csv
import io
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_DIR = ROOT / "outputs" / "md"
DEFAULT_OUTPUT_DIR = ROOT / "outputs" / "entry_exit_2"

SLUG_LEVEL_RE = re.compile(r"(?<![A-Za-z0-9])level-\d+(?:-\d+)*", re.IGNORECASE)
PLAIN_LEVEL_RE = re.compile(r"(?<![A-Za-z0-9])[Ll]evel\s+(\d+(?:\.\d+)*)")
HEADING_RE = re.compile(r"^(#{1,6})\s*(.+?)\s*$", re.MULTILINE)
AUTHOR_SECTION_RE = re.compile(
    r"^#{1,4}\s*(?:作者|授权|脚注|来源|图源|译者|翻译|版权|协议|附件|附录|投票|评论).*$",
    re.MULTILINE,
)

KW_PATH = ROOT / "config" / "kw.json"

def _load_plan2_keywords():
    k = json.loads(KW_PATH.read_text(encoding="utf-8"))
    kw = k["plan2"]
    kw["entrance_aliases"] = k["entrance_aliases"]
    kw["exit_aliases"] = k["exit_aliases"]
    kw["combined_aliases"] = k["combined_aliases"]
    g = globals()
    g["ENTRANCE_ALIASES"] = set(kw["entrance_aliases"])
    g["EXIT_ALIASES"] = set(kw["exit_aliases"])
    g["COMBINED_ALIASES"] = set(kw["combined_aliases"])
    g["ENTRANCE_VERBS"] = set(kw["entrance_verbs"])
    g["EXIT_VERBS"] = set(kw["exit_verbs"])
    g["IRRELEVANT_VERBS"] = set(kw["irrelevant_verbs"])
    g["HEADING_RE"] = re.compile(kw["heading_re"], re.MULTILINE)
    g["SLUG_LEVEL_RE"] = re.compile(kw["slug_level_re"], re.IGNORECASE)
    g["PLAIN_LEVEL_RE"] = re.compile(kw["plain_level_re"])
    g["AUTHOR_SECTION_RE"] = re.compile(kw["author_section_re"], re.MULTILINE)

_load_plan2_keywords()

def strip_front_matter(text: str) -> str:
    lines = text.splitlines()
    if len(lines) >= 3 and lines[0].strip() == "---":
        for index in range(1, len(lines)):
            if lines[index].strip() == "---":
                return "\n".join(lines[index + 1 :])
    return text


def fm_line_count(text: str) -> int:
    lines = text.splitlines()
    if len(lines) >= 3 and lines[0].strip() == "---":
        for i in range(1, len(lines)):
            if lines[i].strip() == "---":
                return i + 1
    return 0


def logical_id_from_filename(path: Path) -> str:
    return path.name.removesuffix(".body.md")


def normalize_level_id(raw: str) -> str:
    lid = raw.lower()
    for prefix in ("trimmed:", "latest:", "old:"):
        if lid.startswith(prefix):
            lid = lid[len(prefix):]
    return lid


def normalize_heading_text(title: str) -> str:
    text = re.sub(r"[*_`~#>]+", "", title)
    text = re.sub(r"[\s:：;；,，.。!！?？()\[\]{}\-_/\\|]+", "", text)
    return text.replace("和", "与")


def heading_type(norm: str) -> str:
    if norm in COMBINED_ALIASES or any(a in norm for a in COMBINED_ALIASES):
        return "combined"
    if norm in ENTRANCE_ALIASES or any(a in norm for a in ENTRANCE_ALIASES):
        return "entrance"
    if norm in EXIT_ALIASES or any(a in norm for a in EXIT_ALIASES):
        return "exit"
    return "other"


def compute_section_intervals(text: str, fm_offset: int) -> dict[str, list[tuple[int, int]]]:
    intervals: dict[str, list[tuple[int, int]]] = {
        "entrance": [], "exit": [], "combined": [], "other": [],
    }
    headings: list[tuple[int, int, str]] = []
    for m in HEADING_RE.finditer(text):
        level = len(m.group(1))
        title = m.group(2).strip()
        norm = normalize_heading_text(title)
        line_no = fm_offset + text[: m.start()].count("\n") + 1
        headings.append((line_no, level, norm))
    if not headings:
        return intervals
    for i, (start_line, h_level, norm) in enumerate(headings):
        htype = heading_type(norm)
        end_line = None
        for j in range(i + 1, len(headings)):
            nxt_line, nxt_level, _ = headings[j]
            if nxt_level <= h_level:
                end_line = nxt_line - 1
                break
        if end_line is None:
            end_line = 99999
        intervals.setdefault(htype, []).append((start_line, end_line))
    return intervals


def extract_all_references(text: str) -> list[tuple[int, str]]:
    refs: list[tuple[int, str]] = []
    for match in SLUG_LEVEL_RE.finditer(text):
        lid = normalize_level_id(match.group(0))
        refs.append((match.start(), lid))
    for match in PLAIN_LEVEL_RE.finditer(text):
        number_part = match.group(1).replace(".", "-")
        lid = f"level-{number_part}".lower()
        refs.append((match.start(), lid))
    refs.sort(key=lambda x: x[0])
    return refs


def deduplicate_refs(refs: list[tuple[int, str]], cluster_dist: int = 40) -> list[tuple[int, str]]:
    if not refs:
        return refs
    result = [refs[0]]
    for pos, target in refs[1:]:
        prev_pos, prev_target = result[-1]
        if target == prev_target and pos - prev_pos <= cluster_dist:
            continue
        result.append((pos, target))
    return result


def char_pos_to_line(text: str, pos: int) -> int:
    return text[:pos].count("\n") + 1


def extract_context(text: str, pos: int, window: int = 100) -> str:
    start = max(0, pos - window)
    end = min(len(text), pos + window)
    return text[start:end].replace("\n", " ").strip()


def classify_edge(context: str, section_type: str) -> tuple[str, str, str]:
    """Return (direction, confidence, reason)."""
    if section_type == "entrance":
        return ("entrance", "high", "在入口标题区间")
    if section_type == "exit":
        return ("exit", "high", "在出口标题区间")

    if section_type == "combined":
        ent_score = sum(1 for v in ENTRANCE_VERBS if re.search(v, context))
        exit_score = sum(1 for v in EXIT_VERBS if re.search(v, context))
        if ent_score > exit_score:
            return ("entrance", "medium", "组合段内方向动词判断(入口>出口)")
        if exit_score > ent_score:
            return ("exit", "medium", "组合段内方向动词判断(出口>入口)")
        return ("irrelevant", "medium", "组合段内无法判定")

    ent_score = sum(1 for v in ENTRANCE_VERBS if re.search(v, context))
    exit_score = sum(1 for v in EXIT_VERBS if re.search(v, context))
    irr_score = sum(1 for v in IRRELEVANT_VERBS if re.search(v, context))

    if irr_score > ent_score + exit_score:
        return ("irrelevant", "low", "无关动词占优")
    if ent_score > exit_score:
        return ("entrance", "low", "句子级方向动词判断(入口>出口)")
    if exit_score > ent_score:
        return ("exit", "low", "句子级方向动词判断(出口>入口)")
    if ent_score > 0 or exit_score > 0:
        if irr_score > 0:
            return ("irrelevant", "low", "含方向词但无关词更多")
        return ("entrance", "low", "微量方向词(偏向入口)")
    return ("irrelevant", "low", "无方向词")


def analyze_file(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")
    logical_id = logical_id_from_filename(path)

    fm_offset = fm_line_count(text)
    full_body = strip_front_matter(text)
    author_cut = AUTHOR_SECTION_RE.search(full_body)
    author_pos = author_cut.start() if author_cut else len(full_body)

    all_refs = extract_all_references(full_body)
    all_refs = deduplicate_refs(all_refs)
    all_refs = [(pos, target) for pos, target in all_refs if pos < author_pos]

    intervals = compute_section_intervals(full_body, fm_offset)

    edges = []
    counts: dict[str, int] = {}
    for pos, target in all_refs:
        if target == logical_id:
            continue
        idx = counts.get(target, 0) + 1
        counts[target] = idx
        line_no = fm_offset + char_pos_to_line(full_body, pos)
        unique_id = f"{logical_id}->{target}#{idx}"
        file_url = Path.cwd() / "outputs" / "md" / path.name
        jump_url = f"file:///{file_url.as_posix()}#L{line_no}"
        ctx = extract_context(full_body, pos)

        sec_type = "other"
        for stype in ("entrance", "exit", "combined", "other"):
            for s_start, s_end in intervals.get(stype, []):
                if s_start <= line_no <= s_end:
                    sec_type = stype
                    break
            if sec_type != "other":
                break

        direction, confidence, reason = classify_edge(ctx, sec_type)

        edges.append({
            "id": unique_id,
            "source": logical_id,
            "target": target,
            "index": idx,
            "position": pos,
            "position_line": line_no,
            "position_url": jump_url,
            "context": ctx[:200],
            "direction": direction,
            "confidence": confidence,
            "reason": reason,
        })

    ent_count = sum(1 for e in edges if e["direction"] == "entrance")
    exit_count = sum(1 for e in edges if e["direction"] == "exit")
    irr_count = sum(1 for e in edges if e["direction"] == "irrelevant")

    return {
        "file": path.name,
        "logical_id": logical_id,
        "total_edges": len(edges),
        "entrance_edges": ent_count,
        "exit_edges": exit_count,
        "irrelevant_edges": irr_count,
        "edges": edges,
    }


def _natural_sort_key(path: Path) -> tuple:
    logical_id = path.name.removesuffix(".body.md")
    parts = logical_id.split("-")
    nums = []
    for p in parts:
        try:
            nums.append(int(p))
        except ValueError:
            nums.append(10**9)
    return (nums, logical_id)


def build_json_output(results: list[dict]) -> dict:
    records = []
    total_edges = 0
    for r in results:
        total_edges += r["total_edges"]
        records.append({
            "source": r["logical_id"],
            "source_file": f"outputs/md/{r['file']}",
            "source_url": f"https://backrooms-wiki-cn.wikidot.com/{r['logical_id']}",
            "summary": {
                "entrance": r["entrance_edges"],
                "exit": r["exit_edges"],
                "irrelevant": r["irrelevant_edges"],
                "total": r["total_edges"],
            },
            "edges": [
                {
                    "id": e["id"],
                    "target": e["target"],
                    "index": e["index"],
                    "position": e["position"],
                    "position_line": e["position_line"],
                    "position_url": e["position_url"],
                    "context": e["context"],
                    "direction": e["direction"],
                    "confidence": e["confidence"],
                    "reason": e["reason"],
                }
                for e in r["edges"]
            ],
        })
    return {"file_count": len(results), "total_edges": total_edges, "records": records}


def build_csv_output(results: list[dict]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow([
        "id", "source", "target", "direction", "confidence", "reason",
        "position", "position_line", "position_url", "context",
    ])
    for r in results:
        for e in r["edges"]:
            writer.writerow([
                e["id"], e["source"], e["target"], e["direction"], e["confidence"],
                e["reason"], e["position"], e["position_line"], e["position_url"],
                e["context"],
            ])
    return output.getvalue()


def build_xlsx_output(results: list[dict], output_path: Path) -> None:
    wb = Workbook()

    # Sheet 1: edges detail
    ws = wb.active
    ws.title = "边分类明细"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    headers = ["id", "source", "target", "direction", "confidence", "reason",
               "position_line", "position_url", "context"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    dir_fills = {
        "entrance": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "exit": PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid"),
        "irrelevant": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }
    dir_fonts = {
        "entrance": Font(color="006100", size=11),
        "exit": Font(color="9C0006", size=11),
        "irrelevant": Font(color="595959", size=11),
    }

    row = 2
    for r in results:
        for e in r["edges"]:
            vals = [e["id"], e["source"], e["target"], e["direction"], e["confidence"],
                    e["reason"], e["position_line"], e["position_url"], e["context"][:150]]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = Font(size=11)
                if col <= 6:
                    cell.alignment = center
            d = e["direction"]
            if d in dir_fills:
                ws.cell(row=row, column=4).fill = dir_fills[d]
                ws.cell(row=row, column=4).font = dir_fonts[d]
            row += 1

    widths = [36, 18, 18, 12, 10, 30, 14, 60, 60]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = f"A1:I{row - 1}"
    ws.freeze_panes = "A2"

    # Sheet 2: per-file summary
    ws2 = wb.create_sheet("文件汇总")
    h2 = [("logical_id", 22), ("总边数", 10), ("入口边", 10), ("出口边", 10), ("无关边", 10)]
    for col, (h, w) in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws2.column_dimensions[get_column_letter(col)].width = w
    for row_idx, r in enumerate(results, 2):
        vals = [r["logical_id"], r["total_edges"], r["entrance_edges"], r["exit_edges"], r["irrelevant_edges"]]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col, value=v)
            cell.alignment = center
            if col == 1:
                cell.font = Font(size=11)
    ws2.auto_filter.ref = f"A1:E{len(results) + 1}"
    ws2.freeze_panes = "A2"

    # Sheet 3: stats
    ws3 = wb.create_sheet("统计")
    total = sum(r["total_edges"] for r in results)
    t_ent = sum(r["entrance_edges"] for r in results)
    t_exit = sum(r["exit_edges"] for r in results)
    t_irr = sum(r["irrelevant_edges"] for r in results)
    high = sum(1 for r in results for e in r["edges"] if e["confidence"] == "high")
    medium = sum(1 for r in results for e in r["edges"] if e["confidence"] == "medium")
    low = sum(1 for r in results for e in r["edges"] if e["confidence"] == "low")

    title_font = Font(bold=True, size=14, color="4472C4")
    ws3.merge_cells("A1:B1")
    ws3.cell(row=1, column=1, value="方案2 边方向分类统计").font = title_font

    stats = [
        ("总边数", total),
        ("入口边", t_ent, f"{t_ent/total*100:.1f}%"),
        ("出口边", t_exit, f"{t_exit/total*100:.1f}%"),
        ("无关边", t_irr, f"{t_irr/total*100:.1f}%"),
        ("", ""),
        ("高置信", high, f"{high/total*100:.1f}%"),
        ("中置信", medium, f"{medium/total*100:.1f}%"),
        ("低置信", low, f"{low/total*100:.1f}%"),
    ]
    for row_idx, stat in enumerate(stats, 2):
        for col, value in enumerate(stat, 1):
            cell = ws3.cell(row=row_idx, column=col, value=value)
            if row_idx == 2:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            else:
                cell.font = Font(size=11)
                cell.alignment = center

    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 16

    wb.save(output_path)


def _short_reason(reason: str) -> str:
    m = {
        "在入口标题区间": "入口标题",
        "在出口标题区间": "出口标题",
        "组合段内无法判定": "组合段",
    }
    if reason in m:
        return m[reason]
    return reason.replace("组合段内方向动词判断(", "组合段(").replace("句子级方向动词判断(", "方向词(").replace("微量方向词(偏向入口)", "弱信号").replace("含方向词但无关词更多", "无关占优").replace("无关动词占优", "无关词").replace("无方向词", "无关")[:12]


def build_summary_json(results: list[dict]) -> dict:
    records = []
    for r in results:
        ent_set: dict[str, str] = {}
        exit_set: dict[str, str] = {}
        irr_set: dict[str, str] = {}
        for e in r["edges"]:
            target = e["target"]
            reason = _short_reason(e["reason"])
            if e["direction"] == "entrance":
                if target not in ent_set:
                    ent_set[target] = reason
            elif e["direction"] == "exit":
                if target not in exit_set:
                    exit_set[target] = reason
            else:
                if target not in irr_set:
                    irr_set[target] = reason
        records.append({
            "source": r["logical_id"],
            "source_url": f"https://backrooms-wiki-cn.wikidot.com/{r['logical_id']}",
            "entrances": [f"{t} ({r})" for t, r in ent_set.items()],
            "exits": [f"{t} ({r})" for t, r in exit_set.items()],
            "irrelevant": [f"{t} ({r})" for t, r in irr_set.items()],
        })
    return {"file_count": len(results), "records": records}


def build_clean_summary_json(results: list[dict]) -> dict:
    """Summary without reason strings — just clean level ID lists."""
    records = []
    for r in results:
        ent_set = set()
        exit_set = set()
        irr_set = set()
        for e in r["edges"]:
            if e["direction"] == "entrance":
                ent_set.add(e["target"])
            elif e["direction"] == "exit":
                exit_set.add(e["target"])
            else:
                irr_set.add(e["target"])
        records.append({
            "source": r["logical_id"],
            "entrances": sorted(ent_set),
            "exits": sorted(exit_set),
            "irrelevant": sorted(irr_set),
        })
    return {"file_count": len(results), "records": records}


def write_outputs(results: list[dict], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "level_references.json"
    summary_path = output_dir / "level_summary.json"
    clean_path = output_dir / "level_summary_clean.json"
    csv_path = output_dir / "level_references.csv"
    xlsx_path = output_dir / "level_references.xlsx"

    json_path.write_text(
        json.dumps(build_json_output(results), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    summary_path.write_text(
        json.dumps(build_summary_json(results), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    clean_path.write_text(
        json.dumps(build_clean_summary_json(results), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    csv_path.write_text(build_csv_output(results), encoding="utf-8")
    build_xlsx_output(results, xlsx_path)

    print(f"Wrote JSON       to {json_path}")
    print(f"Wrote Summary    to {summary_path}")
    print(f"Wrote Clean      to {clean_path}")
    print(f"Wrote CSV        to {csv_path}")
    print(f"Wrote XLSX       to {xlsx_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="方案2: Extract Level references and classify direction.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--pattern", default="*.body.md")
    parser.add_argument("--limit", type=int)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    files = sorted(args.input_dir.glob(args.pattern), key=_natural_sort_key)
    if args.limit is not None:
        files = files[: args.limit]
    results = [analyze_file(path) for path in files]
    write_outputs(results, args.output_dir)


if __name__ == "__main__":
    main()