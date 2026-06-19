from __future__ import annotations

import argparse
import csv
import io
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_DIR = ROOT / "outputs" / "md"
DEFAULT_OUTPUT_DIR = ROOT / "outputs" / "entry_exit"

HEADING_RE = re.compile(r"^(#{1,6})\s*(.+?)\s*$")
LINK_RE = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")
IMAGE_RE = re.compile(r"!\[([^\]]*)\]\([^)]+\)")
EMPHASIS_RE = re.compile(r"[*_`~#>]+")
PUNCT_RE = re.compile(r"[\s:：;；,，.。!！?？()\[\]{}\-_/\\|]+")

PLAIN_LEVEL_RE = re.compile(r"(?<![A-Za-z0-9])[Ll]evel\s*(-?\d+(?:\.\d+)*)")
SLUG_LEVEL_RE = re.compile(r"(?<![A-Za-z0-9])level-\d+(?:-\d+)*", re.IGNORECASE)
PATH_LEVEL_RE = re.compile(r"(?:^|/)(?:trimmed:|latest:)?(level-\d+(?:-\d+)*)$", re.IGNORECASE)

ENTRANCE_ALIASES = {
    "入口",
    "如何进入",
    "进入方式",
    "进入方法",
    "进入途径",
}

EXIT_ALIASES = {
    "出口",
    "如何离开",
    "离开方式",
    "离开方法",
    "离开途径",
}

COMBINED_ALIASES = {
    "入口与出口",
    "入口和出口",
    "出入口",
    "入口出口",
}

ENTRY_EXIT_SIGNAL_RE = re.compile(r"(入口|出口|进入|离开|逃离|切出|返回|通往|通过|经由)")

ENTRANCE_KEYWORDS = (
    "进入",
    "进入到",
    "抵达",
    "通往本层级",
    "进入本层级",
    "入口",
    "可经由",
    "可以从",
    "通过",
)

EXIT_KEYWORDS = (
    "离开",
    "逃离",
    "出去",
    "通往",
    "出口",
    "返回",
    "切出",
    "原路返回",
)


@dataclass
class Section:
    title: str
    level: int
    lines: list[str] = field(default_factory=list)
    children: list["Section"] = field(default_factory=list)


def strip_front_matter(text: str) -> str:
    lines = text.splitlines()
    if len(lines) >= 3 and lines[0].strip() == "---":
        for index in range(1, len(lines)):
            if lines[index].strip() == "---":
                return "\n".join(lines[index + 1 :])
    return text


def normalize_heading(text: str) -> str:
    text = LINK_RE.sub(r"\1", text)
    text = EMPHASIS_RE.sub("", text)
    text = text.strip()
    text = PUNCT_RE.sub("", text)
    text = text.replace("和", "与")
    return text


def parse_markdown_sections(text: str) -> Section:
    root = Section(title="", level=0)
    stack: list[Section] = [root]

    for raw_line in strip_front_matter(text).splitlines():
        line = raw_line.rstrip()
        heading_match = HEADING_RE.match(line)
        if heading_match:
            hashes, title = heading_match.groups()
            level = len(hashes)
            section = Section(title=title.strip(), level=level)
            while stack and stack[-1].level >= level:
                stack.pop()
            stack[-1].children.append(section)
            stack.append(section)
            continue

        stack[-1].lines.append(line)

    return root


def collect_text_lines(section: Section) -> list[str]:
    lines = list(section.lines)
    for child in section.children:
        lines.extend(collect_text_lines(child))
    return lines


def clean_block(block: str) -> str:
    block = IMAGE_RE.sub(r"\1", block)
    block = LINK_RE.sub(r"\1", block)
    block = EMPHASIS_RE.sub("", block)
    block = re.sub(r"\s+", " ", block).strip()
    return block


def split_blocks(lines: Iterable[str]) -> list[str]:
    blocks: list[str] = []
    buffer: list[str] = []

    def flush() -> None:
        if not buffer:
            return
        text = clean_block(" ".join(buffer))
        if text:
            blocks.append(text)
        buffer.clear()

    for line in lines:
        stripped = line.strip()
        if not stripped:
            flush()
            continue

        if stripped.startswith(("-", "*")) or re.match(r"^\d+\.", stripped):
            flush()
            text = clean_block(re.sub(r"^(-|\*|\d+\.)\s*", "", stripped))
            if text:
                blocks.append(text)
            continue

        buffer.append(stripped)

    flush()
    return blocks


def classify_block(block: str) -> str | None:
    if not ENTRY_EXIT_SIGNAL_RE.search(block):
        return None

    entrance_score = sum(keyword in block for keyword in ENTRANCE_KEYWORDS)
    exit_score = sum(keyword in block for keyword in EXIT_KEYWORDS)

    if "离开" in block or "逃离" in block or "原路返回" in block:
        exit_score += 2
    if "进入" in block and "离开" not in block and "逃离" not in block:
        entrance_score += 1

    if entrance_score == 0 and exit_score == 0:
        return None
    if entrance_score > exit_score:
        return "entrances"
    if exit_score > entrance_score:
        return "exits"

    if "进入" in block and "离开" not in block:
        return "entrances"
    if "离开" in block and "进入" not in block:
        return "exits"
    return None


def find_sections(section: Section, normalized_titles: set[str]) -> list[Section]:
    matches: list[Section] = []
    for child in section.children:
        if normalize_heading(child.title) in normalized_titles:
            matches.append(child)
        matches.extend(find_sections(child, normalized_titles))
    return matches


def unique_blocks(blocks: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for block in blocks:
        if block and block not in seen:
            seen.add(block)
            result.append(block)
    return result


def extract_from_combined_section(section: Section) -> tuple[list[str], list[str]]:
    entrances: list[str] = []
    exits: list[str] = []

    direct_entrances = [child for child in section.children if normalize_heading(child.title) in ENTRANCE_ALIASES]
    direct_exits = [child for child in section.children if normalize_heading(child.title) in EXIT_ALIASES]

    for child in direct_entrances:
        entrances.extend(split_blocks(collect_text_lines(child)))
    for child in direct_exits:
        exits.extend(split_blocks(collect_text_lines(child)))

    if entrances or exits:
        return unique_blocks(entrances), unique_blocks(exits)

    current_bucket: str | None = None
    for block in split_blocks(section.lines):
        normalized_block = normalize_heading(block)
        if normalized_block in ENTRANCE_ALIASES:
            current_bucket = "entrances"
            continue
        if normalized_block in EXIT_ALIASES:
            current_bucket = "exits"
            continue

        if current_bucket == "entrances":
            entrances.append(block)
            continue
        if current_bucket == "exits":
            exits.append(block)
            continue

        bucket = classify_block(block)
        if bucket == "entrances":
            entrances.append(block)
        elif bucket == "exits":
            exits.append(block)

    return unique_blocks(entrances), unique_blocks(exits)


def extract_entry_exit(text: str) -> dict[str, list[str]]:
    tree = parse_markdown_sections(text)
    entrances: list[str] = []
    exits: list[str] = []

    combined_sections = find_sections(tree, COMBINED_ALIASES)
    for section in combined_sections:
        combined_entrances, combined_exits = extract_from_combined_section(section)
        entrances.extend(combined_entrances)
        exits.extend(combined_exits)

    standalone_entrances = find_sections(tree, ENTRANCE_ALIASES)
    standalone_exits = find_sections(tree, EXIT_ALIASES)

    for section in standalone_entrances:
        entrances.extend(split_blocks(section.lines))

    for section in standalone_exits:
        exits.extend(split_blocks(section.lines))

    return {
        "entrances": unique_blocks(entrances),
        "exits": unique_blocks(exits),
    }


def logical_id_from_filename(path: Path) -> str:
    return path.name.removesuffix(".body.md")


def normalize_level_token(token: str) -> str:
    value = token.lower().strip()
    if value.startswith("level-"):
        return value
    if value.startswith("level "):
        value = value.replace("level ", "level-", 1)
    if value.startswith("level"):
        value = value.replace("level", "level-", 1)
    value = value.replace(".", "-")
    value = re.sub(r"-{2,}", "-", value)
    return value


def extract_level_ids_from_text(text: str) -> list[str]:
    level_ids: list[str] = []

    for match in SLUG_LEVEL_RE.finditer(text):
        level_ids.append(match.group(0).lower())

    for match in PLAIN_LEVEL_RE.finditer(text):
        number_part = match.group(1).replace(".", "-")
        level_ids.append(f"level-{number_part}".lower())

    return unique_blocks(level_ids)


def extract_level_ids_from_block(block: str, current_level_id: str) -> list[str]:
    level_ids = [level_id for level_id in extract_level_ids_from_text(block) if level_id != current_level_id]
    return unique_blocks(level_ids)


def analyze_file(path: Path) -> dict[str, object]:
    text = path.read_text(encoding="utf-8")
    extracted = extract_entry_exit(text)
    logical_id = logical_id_from_filename(path)

    entrance_level_ids = unique_blocks(
        level_id
        for block in extracted["entrances"]
        for level_id in extract_level_ids_from_block(block, logical_id)
    )
    exit_level_ids = unique_blocks(
        level_id
        for block in extracted["exits"]
        for level_id in extract_level_ids_from_block(block, logical_id)
    )

    return {
        "file": path.name,
        "logical_id": logical_id,
        "entrance_count": len(extracted["entrances"]),
        "exit_count": len(extracted["exits"]),
        "entrances": extracted["entrances"],
        "exits": extracted["exits"],
        "entrance_level_ids": entrance_level_ids,
        "exit_level_ids": exit_level_ids,
        "missing_entrance_level_ids": bool(extracted["entrances"]) and not entrance_level_ids,
        "missing_exit_level_ids": bool(extracted["exits"]) and not exit_level_ids,
    }


def build_description_output(results: list[dict[str, object]]) -> dict[str, object]:
    return {
        "file_count": len(results),
        "with_entrances": sum(bool(item["entrances"]) for item in results),
        "with_exits": sum(bool(item["exits"]) for item in results),
        "results": results,
    }


def build_level_output(results: list[dict[str, object]]) -> dict[str, object]:
    compact_results = []
    for item in results:
        compact_results.append(
            {
                "file": item["file"],
                "logical_id": item["logical_id"],
                "entrance_level_ids": item["entrance_level_ids"],
                "exit_level_ids": item["exit_level_ids"],
            }
        )

    return {
        "file_count": len(compact_results),
        "with_entrance_level_ids": sum(bool(item["entrance_level_ids"]) for item in compact_results),
        "with_exit_level_ids": sum(bool(item["exit_level_ids"]) for item in compact_results),
        "results": compact_results,
    }


def build_missing_level_info_text(results: list[dict[str, object]]) -> str:
    lines = ["Missing entrance/exit level ids", ""]

    for item in results:
        missing_entry = item["missing_entrance_level_ids"]
        missing_exit = item["missing_exit_level_ids"]
        if not missing_entry and not missing_exit:
            continue

        lines.append(f"{item['logical_id']} ({item['file']})")
        if missing_entry:
            lines.append("  [entrances]")
            for block in item["entrances"]:
                lines.append(f"  - {block}")
        if missing_exit:
            lines.append("  [exits]")
            for block in item["exits"]:
                lines.append(f"  - {block}")
        lines.append("")

    if len(lines) == 2:
        lines.append("None")

    return "\n".join(lines)


def build_missing_descriptions_text(results: list[dict[str, object]]) -> str:
    lines = ["Missing entrance/exit descriptions", ""]

    for item in results:
        missing_entrances = not item["entrances"]
        missing_exits = not item["exits"]
        if not missing_entrances and not missing_exits:
            continue

        lines.append(f"{item['logical_id']} ({item['file']})")
        if missing_entrances:
            lines.append("  - missing entrances")
        if missing_exits:
            lines.append("  - missing exits")
        lines.append("")

    if len(lines) == 2:
        lines.append("None")

    return "\n".join(lines)


def build_status_report_csv(results: list[dict[str, object]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["logical_id", "入口描述", "入口层级", "出口描述", "出口层级"])

    for item in results:
        has_entrance_description = bool(item["entrances"])
        has_entrance_level_ids = bool(item["entrance_level_ids"])
        has_exit_description = bool(item["exits"])
        has_exit_level_ids = bool(item["exit_level_ids"])

        writer.writerow([
            item["logical_id"],
            has_entrance_description,
            has_entrance_level_ids,
            has_exit_description,
            has_exit_level_ids,
        ])

    return output.getvalue()


def build_status_report_xlsx(results: list[dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "出入口状态"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font = Font(color="006100", size=11)
    red_font = Font(color="9C0006", size=11)

    headers = ["logical_id", "入口描述", "入口层级", "出口描述", "出口层级"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx, item in enumerate(results, 2):
        has_entrance_description = bool(item["entrances"])
        has_entrance_level_ids = bool(item["entrance_level_ids"])
        has_exit_description = bool(item["exits"])
        has_exit_level_ids = bool(item["exit_level_ids"])

        values = [
            item["logical_id"],
            "✔ 有" if has_entrance_description else "✘ 无",
            "✔ 有" if has_entrance_level_ids else "✘ 无",
            "✔ 有" if has_exit_description else "✘ 无",
            "✔ 有" if has_exit_level_ids else "✘ 无",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = center_align
            if col == 1:
                cell.font = Font(size=11)
            elif "✔" in str(value):
                cell.fill = green_fill
                cell.font = green_font
            else:
                cell.fill = red_fill
                cell.font = red_font

    col_widths = [24, 12, 12, 12, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = f"A1:E{len(results) + 1}"
    ws.freeze_panes = "A2"

    # ----- Sheet 2: 统计分析 -----
    ws2 = wb.create_sheet("统计分析")

    total = len(results)
    ent_desc = sum(1 for r in results if r["entrances"])
    ent_desc_id = sum(1 for r in results if r["entrances"] and r["entrance_level_ids"])
    exit_desc = sum(1 for r in results if r["exits"])
    exit_desc_id = sum(1 for r in results if r["exits"] and r["exit_level_ids"])
    both_desc = sum(1 for r in results if r["entrances"] and r["exits"])
    both_desc_id = sum(1 for r in results if r["entrances"] and r["entrance_level_ids"] and r["exits"] and r["exit_level_ids"])

    stats = [
        ("总文章数", total, ""),
        ("", "", ""),
        ("指标", "命中数", "成功率"),
        ("入口描述文本", ent_desc, f"{ent_desc/total*100:.1f}%"),
        ("入口描述文本 + 层级ID", ent_desc_id, f"{ent_desc_id/total*100:.1f}%"),
        ("出口描述文本", exit_desc, f"{exit_desc/total*100:.1f}%"),
        ("出口描述文本 + 层级ID", exit_desc_id, f"{exit_desc_id/total*100:.1f}%"),
        ("入口 + 出口描述文本", both_desc, f"{both_desc/total*100:.1f}%"),
        ("入口 + 出口描述文本 + 双ID", both_desc_id, f"{both_desc_id/total*100:.1f}%"),
    ]

    title_font = Font(bold=True, size=14, color="4472C4")
    header_fill2 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white2 = Font(bold=True, size=11, color="FFFFFF")
    percent_font_green = Font(bold=True, size=11, color="006100")
    percent_font_yellow = Font(bold=True, size=11, color="9C6500")
    percent_font_red = Font(bold=True, size=11, color="9C0006")

    ws2.merge_cells("A1:C1")
    ws2.cell(row=1, column=1, value="出入口信息获取成功率统计").font = title_font

    for row_idx, (label, count, rate) in enumerate(stats, 2):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=count if count != "" else "")
        cell_c = ws2.cell(row=row_idx, column=3, value=rate)

        if label == "指标":
            for c in [cell_a, cell_b, cell_c]:
                c.font = header_font_white2
                c.fill = header_fill2
                c.alignment = center_align
        elif label and not label.startswith("总"):
            cell_a.alignment = Alignment(horizontal="right")
            cell_b.alignment = center_align
            cell_c.alignment = center_align

            pct = float(rate.strip("%"))
            if pct >= 60:
                cell_c.font = percent_font_green
            elif pct >= 45:
                cell_c.font = percent_font_yellow
            else:
                cell_c.font = percent_font_red

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    wb.save(output_path)


def write_outputs(results: list[dict[str, object]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    description_path = output_dir / "entry_exit_descriptions.json"
    level_path = output_dir / "entry_exit_level_ids.json"
    status_csv_path = output_dir / "entry_exit_status.csv"
    status_xlsx_path = output_dir / "entry_exit_status.xlsx"

    description_path.write_text(
        json.dumps(build_description_output(results), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    level_path.write_text(
        json.dumps(build_level_output(results), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    status_csv_path.write_text(build_status_report_csv(results), encoding="utf-8")
    build_status_report_xlsx(results, status_xlsx_path)

    for stale_txt in [
        output_dir / "missing_level_ids.txt",
        output_dir / "missing_entry_exit_descriptions.txt",
    ]:
        if stale_txt.exists():
            stale_txt.unlink()

    print(f"Wrote description output to {description_path}")
    print(f"Wrote level-id output to {level_path}")
    print(f"Wrote status report to {status_csv_path}")
    print(f"Wrote status xlsx to {status_xlsx_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract Backrooms entry and exit descriptions and level ids.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR, help="Directory containing markdown files.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for generated output files.")
    parser.add_argument("--pattern", default="*.body.md", help="Glob pattern used to select files.")
    parser.add_argument("--limit", type=int, help="Optional limit for debugging.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    files = sorted(args.input_dir.glob(args.pattern))
    if args.limit is not None:
        files = files[: args.limit]

    results = [analyze_file(path) for path in files]
    write_outputs(results, args.output_dir)


if __name__ == "__main__":
    main()
