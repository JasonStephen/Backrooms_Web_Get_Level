from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
PLAN1_PATH = ROOT / "outputs" / "entry_exit_1" / "entry_exit_level_ids.json"
PLAN2_PATH = ROOT / "outputs" / "entry_exit_2" / "level_summary_clean.json"
OUTPUT_DIR = ROOT / "outputs" / "entry_exit_compare"


def _natural_sort_key(key: str) -> tuple:
    parts = key.split("-")
    nums = []
    for p in parts:
        try:
            nums.append(int(p))
        except ValueError:
            nums.append(10**9)
    return (nums, key)


def compare_results():
    with open(PLAN1_PATH, "r", encoding="utf-8") as f:
        p1_data = json.load(f)
    with open(PLAN2_PATH, "r", encoding="utf-8") as f:
        p2_data = json.load(f)

    p1_map = {r["logical_id"]: r for r in p1_data["results"]}
    p2_map = {r["source"]: r for r in p2_data["records"]}

    results = []
    all_sources = sorted(set(p1_map.keys()) | set(p2_map.keys()), key=_natural_sort_key)

    total_ent_agree = total_ent_p1only = total_ent_p2only = 0
    total_exit_agree = total_exit_p1only = total_exit_p2only = 0

    for src in all_sources:
        p1_ent = set(p1_map[src]["entrance_level_ids"]) if src in p1_map else set()
        p1_exit = set(p1_map[src]["exit_level_ids"]) if src in p1_map else set()
        p2_ent = set(p2_map[src]["entrances"]) if src in p2_map else set()
        p2_exit = set(p2_map[src]["exits"]) if src in p2_map else set()

        ent_agree = sorted(p1_ent & p2_ent)
        ent_p1only = sorted(p1_ent - p2_ent)
        ent_p2only = sorted(p2_ent - p1_ent)
        exit_agree = sorted(p1_exit & p2_exit)
        exit_p1only = sorted(p1_exit - p2_exit)
        exit_p2only = sorted(p2_exit - p1_exit)

        total_ent_agree += len(ent_agree)
        total_ent_p1only += len(ent_p1only)
        total_ent_p2only += len(ent_p2only)
        total_exit_agree += len(exit_agree)
        total_exit_p1only += len(exit_p1only)
        total_exit_p2only += len(exit_p2only)

        results.append({
            "source": src,
            "entrance_agree": ent_agree,
            "entrance_plan1_only": ent_p1only,
            "entrance_plan2_only": ent_p2only,
            "exit_agree": exit_agree,
            "exit_plan1_only": exit_p1only,
            "exit_plan2_only": exit_p2only,
        })

    stats = {
        "total_sources": len(all_sources),
        "entrance": {
            "agree": total_ent_agree,
            "plan1_only": total_ent_p1only,
            "plan2_only": total_ent_p2only,
            "agree_pct": total_ent_agree / max(total_ent_agree + total_ent_p1only + total_ent_p2only, 1) * 100,
        },
        "exit": {
            "agree": total_exit_agree,
            "plan1_only": total_exit_p1only,
            "plan2_only": total_exit_p2only,
            "agree_pct": total_exit_agree / max(total_exit_agree + total_exit_p1only + total_exit_p2only, 1) * 100,
        },
    }
    return results, stats


def write_outputs(results, stats, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)

    json_data = {"stats": stats, "records": results}
    json_path = output_dir / "plan_comparison.json"
    json_path.write_text(json.dumps(json_data, ensure_ascii=False, indent=2), encoding="utf-8")

    xlsx_path = output_dir / "plan_comparison.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "逐层对比"
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfn = Font(bold=True, size=11, color="FFFFFF")
    ca = Alignment(horizontal="center", vertical="center")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    headers = [
        "source", "入口一致", "入口仅方案1", "入口仅方案2",
        "出口一致", "出口仅方案1", "出口仅方案2",
        "入一致", "入P1多", "入P2多", "出一致", "出P1多", "出P2多",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfn; c.fill = hf; c.alignment = ca

    row = 2
    for r in results:
        has_diff = (r["entrance_plan1_only"] or r["entrance_plan2_only"] or
                    r["exit_plan1_only"] or r["exit_plan2_only"])
        vals = [
            r["source"],
            " | ".join(r["entrance_agree"]),
            " | ".join(r["entrance_plan1_only"]),
            " | ".join(r["entrance_plan2_only"]),
            " | ".join(r["exit_agree"]),
            " | ".join(r["exit_plan1_only"]),
            " | ".join(r["exit_plan2_only"]),
            len(r["entrance_agree"]), len(r["entrance_plan1_only"]), len(r["entrance_plan2_only"]),
            len(r["exit_agree"]), len(r["exit_plan1_only"]), len(r["exit_plan2_only"]),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(size=11)
            cell.alignment = ca if col <= 1 or col >= 8 else Alignment(horizontal="left", vertical="center")
        if has_diff:
            ws.cell(row=row, column=1).fill = yellow
        else:
            ws.cell(row=row, column=1).fill = green
        row += 1

    widths = [22, 18, 18, 18, 18, 18, 18, 8, 8, 8, 8, 8, 8]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = f"A1:M{row - 1}"
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("统计")
    tf = Font(bold=True, size=14, color="4472C4")
    ws2.merge_cells("A1:C1")
    ws2.cell(row=1, column=1, value="方案1 vs 方案2 对比统计").font = tf

    srows = [
        ("", "", ""),
        ("指标", "数量", "占比"),
        ("总层级数", stats["total_sources"], ""),
        ("", "", ""),
        ("入口 — 两方案一致", stats["entrance"]["agree"], f"{stats['entrance']['agree_pct']:.1f}%"),
        ("入口 — 仅方案1有", stats["entrance"]["plan1_only"], ""),
        ("入口 — 仅方案2有", stats["entrance"]["plan2_only"], ""),
        ("", "", ""),
        ("出口 — 两方案一致", stats["exit"]["agree"], f"{stats['exit']['agree_pct']:.1f}%"),
        ("出口 — 仅方案1有", stats["exit"]["plan1_only"], ""),
        ("出口 — 仅方案2有", stats["exit"]["plan2_only"], ""),
    ]
    for ri, stat in enumerate(srows, 2):
        for ci, v in enumerate(stat, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = hfn if ri == 3 else Font(size=11)
            cell.alignment = ca
            if ri == 3:
                cell.fill = hf

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    wb.save(xlsx_path)
    print(f"Wrote JSON to {json_path}")
    print(f"Wrote XLSX to {xlsx_path}")


def main():
    results, stats = compare_results()
    write_outputs(results, stats, OUTPUT_DIR)
    print()
    es = stats["entrance"]
    xs = stats["exit"]
    print(f"Entrance: agree={es['agree']} ({es['agree_pct']:.1f}%), p1_only={es['plan1_only']}, p2_only={es['plan2_only']}")
    print(f"Exit:     agree={xs['agree']} ({xs['agree_pct']:.1f}%), p1_only={xs['plan1_only']}, p2_only={xs['plan2_only']}")


if __name__ == "__main__":
    main()
